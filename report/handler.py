import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelHandler:
    @staticmethod
    def read_file(filename: str, sheet_name: str = "Заказы_сырье") -> list[dict]:
        logger.info("Чтение файла '%s', лист '%s'", filename, sheet_name)
        wb = load_workbook(filename=filename, read_only=True)
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        logger.debug("Заголовки: %s", headers)
        orders = []
        for row, value in enumerate(rows, start=2):
            data = dict(zip(headers, value))
            data["Номер строки"] = row
            orders.append(data)

        wb.close()
        logger.info("Прочитано %d строк", len(orders))
        return orders
