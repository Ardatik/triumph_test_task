from collections import defaultdict
from decimal import Decimal

from openpyxl import Workbook

from .handler import ExcelHandler
from .models import ReportOrder, ValidationError
from .normalization import Normalization
from .validation import OrderValidator


class ReportBuilder:
    _HEADERS = [
        "Мастер",
        "Количество выполненных заказов",
        "Сумма выполненных заказов",
        "Количество рекламаций",
        "% рекламаций",
        "Количество заказов без фото",
        "Количество отмененных заказов",
    ]

    def __init__(self):
        self.handler = ExcelHandler()
        self.normalizer = Normalization()
        self.validator = OrderValidator(self.normalizer)

    def build_report(self, input_file: str, output_file: str = "report.xlsx") -> None:
        raw_data = self.handler.read_file(filename=input_file)
        valid_orders, errors = self.validator.validate(data=raw_data)
        master_stats = self._prepare_data_for_report(valid_orders)
        self._write_report_to_file(master_stats, errors, output_file)

    def _prepare_data_for_report(self, orders: list[ReportOrder]) -> dict[str, dict]:
        master_stats = defaultdict(
            lambda: {
                "completed": 0,
                "sum_completed": Decimal(0),
                "complaints": 0,
                "canceled": 0,
                "without_photo": 0,
            }
        )
        for order in orders:
            stats = master_stats[order.master]
            if order.status == "Выполнен":
                stats["completed"] += 1
                stats["sum_completed"] += order.amount
                if order.complaint:
                    stats["complaints"] += 1
                if not order.photo:
                    stats["without_photo"] += 1
            elif order.status == "Отменен":
                stats["canceled"] += 1
        return master_stats

    def _write_report_to_file(
        self,
        master_stats: dict[str, dict],
        errors: list[ValidationError],
        output_file: str,
    ) -> None:
        wb = Workbook()
        ws_report = wb.create_sheet("Отчёт по мастерам", 0)
        for col, header in enumerate(self._HEADERS, 1):
            ws_report.cell(row=1, column=col, value=header)
        row = 2
        for master, stats in sorted(master_stats.items()):
            completed = stats["completed"]
            complaints = stats["complaints"]
            percent = round(complaints / completed * 100, 2) if completed else 0.0
            ws_report.cell(row=row, column=1, value=master)
            ws_report.cell(row=row, column=2, value=completed)
            ws_report.cell(row=row, column=3, value=float(stats["sum_completed"]))
            ws_report.cell(row=row, column=4, value=complaints)
            ws_report.cell(row=row, column=5, value=percent)
            ws_report.cell(row=row, column=6, value=stats["without_photo"])
            ws_report.cell(row=row, column=7, value=stats["canceled"])
            row += 1
        for col in range(1, len(self._HEADERS) + 1):
            column_letter = ws_report.cell(row=1, column=col).column_letter
            ws_report.column_dimensions[column_letter].width = 25
        ws_errors = wb.create_sheet("Ошибки")
        error_headers = ["Номер строки", "ID заказа", "Поле", "Сообщение", "Значение"]
        for col, header in enumerate(error_headers, 1):
            ws_errors.cell(row=1, column=col, value=header)
        for i, err in enumerate(errors, 2):
            ws_errors.cell(row=i, column=1, value=err.row_number)
            ws_errors.cell(row=i, column=2, value=err.order_id)
            ws_errors.cell(row=i, column=3, value=err.field)
            ws_errors.cell(row=i, column=4, value=err.message)
            if err.value is not None:
                ws_errors.cell(row=i, column=5, value=str(err.value))
            else:
                ws_errors.cell(row=i, column=5, value="")
        for col in range(1, len(error_headers) + 1):
            column_letter = ws_errors.cell(row=1, column=col).column_letter
            ws_errors.column_dimensions[column_letter].width = 25
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(output_file)
        print(f"Отчёт сохранён в {output_file}")
